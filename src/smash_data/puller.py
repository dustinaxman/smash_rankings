import requests
import os
import time
from datetime import datetime
import pandas as pd
import boto3
import re
from openpyxl import load_workbook
from botocore.exceptions import ClientError
from boto3.dynamodb.conditions import Key
import json
import logging
from backoff import on_exception, expo
from ratelimit import limits, RateLimitException
from src.utils.constants import tier_mapper, dynamo_db_table_name, s3_bucket, LOG_FOLDER_PATH
import subprocess

os.makedirs(LOG_FOLDER_PATH, exist_ok=True)
log_file_name = f"log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
log_file_path = os.path.join(LOG_FOLDER_PATH, log_file_name)

logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    handlers=[
                        logging.StreamHandler(),                    # Console handler
                        logging.FileHandler(log_file_path, mode='a')  # File handler with append mode
                    ])


def get_parameter_from_ssm(parameter_name, region="us-east-1"):
    ssm = boto3.client('ssm', region_name=region)
    try:
        response = ssm.get_parameter(Name=parameter_name, WithDecryption=True)
        return response['Parameter']['Value']
    except ClientError as e:
        logging.info(f"Error fetching parameter {parameter_name}: {e}")
        return None

if not os.environ.get("STARTGG_API_KEY"):
    startgg_api_key = get_parameter_from_ssm("STARTGG_API_KEY")
    if startgg_api_key:
        os.environ["STARTGG_API_KEY"] = startgg_api_key

token = os.environ.get("STARTGG_API_KEY")

s3 = boto3.client('s3')
dynamodb = boto3.resource('dynamodb')




get_events_query = """
query TournamentEvents($tourneySlug:String, $videogameId:[ID]!) {
  tournament(slug: $tourneySlug) {
    id
    name
    events(filter:{videogameId: $videogameId}) {
      id
      name
    }
  }
}
"""


get_event_id_from_slug_query = """
query getEventId($slug: String) {
  event(slug: $slug) {
    id
    name
  }
}
"""

@limits(calls=80, period=60)
@on_exception(expo, RateLimitException, max_tries=5)
@on_exception(expo, Exception, max_tries=5)
def get_event_id(overall_slug, token):
    url = "https://api.start.gg/gql/alpha"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }

    variables = {"slug": overall_slug}

    response = requests.post(url, json={'query': get_event_id_from_slug_query, 'variables': variables}, headers=headers)
    if response.status_code == 200:
        data = response.json()
        if "data" in data and "event" in data["data"] and "id" in data["data"]["event"]:
            return data["data"]["event"]["id"]
        else:
            raise Exception("Invalid response structure or missing data")
    else:
        raise Exception(f"Request failed with status {response.status_code}")


def weighted_edit_distance(s1, s2, digit_weight):
    """
    Calculate a weighted edit distance where deletions and insertions of digits have a higher cost.
    """
    len1, len2 = len(s1), len(s2)
    # Initialize DP table
    dp = [[0]*(len2+1) for _ in range(len1+1)]

    # Initialize first column and first row of the DP table
    for i in range(1, len1+1):
        dp[i][0] = dp[i-1][0] + (digit_weight if s1[i-1].isdigit() else 1)
    for j in range(1, len2+1):
        dp[0][j] = dp[0][j-1] + (digit_weight if s2[j-1].isdigit() else 1)

    # Fill DP table
    for i in range(1, len1+1):
        for j in range(1, len2+1):
            if s1[i-1] == s2[j-1]:
                cost_sub = dp[i-1][j-1]  # No cost if characters are the same
            else:
                cost_sub = dp[i-1][j-1] + 1  # Substitution cost

            cost_del = dp[i-1][j] + (digit_weight if s1[i-1].isdigit() else 1)
            cost_ins = dp[i][j-1] + (digit_weight if s2[j-1].isdigit() else 1)

            dp[i][j] = min(cost_sub, cost_del, cost_ins)

    return dp[len1][len2]

def find_closest_strings(list1, list2, digit_weight=1.5):
    """
    For each string in list1, find the string in list2 with the minimum weighted edit distance.
    """
    closest_strings = []
    for s1 in list1:
        closest_match = min(list2, key=lambda s2: weighted_edit_distance(s1, s2, digit_weight))
        closest_strings.append(closest_match)
    return closest_strings


@limits(calls=80, period=60)
@on_exception(expo, RateLimitException, max_tries=5)
@on_exception(expo, Exception, max_tries=5)
def get_tournament_events(tourney_slug, token, videogame_id=1386):
    url = "https://api.start.gg/gql/alpha"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }

    variables = {
        "tourneySlug": tourney_slug,
        "videogameId": [videogame_id]
    }

    response = requests.post(url, json={'query': get_events_query, 'variables': variables}, headers=headers)
    if response.status_code == 200:
        data = response.json()
        if "data" in data and "tournament" in data["data"] and "events" in data["data"]["tournament"]:
            tournament_name = data["data"]["tournament"]["name"]
            return tournament_name, [{"name": event["name"].lower(), "id": event["id"]} for event in data["data"]["tournament"]["events"]]
        else:
            raise Exception("get_tournament_events Invalid response structure or missing data")
    else:
        raise Exception(f"get_tournament_events Request failed with status {response.status_code}")

def select_correct_event(all_events_for_tourney):
    exclude_keywords = [
        "double", "squad", "team", "crew", "2v2", "random", "low tier", 
        "speedrun", "ladder", "lcq", "dobles", "16 and under", "n poke", 
        "vip", "all stars", "swiss gp"
    ]
    include_keywords = ["smash", "singles", "ultimate", "1v1", "1vs1"]
    all_selections = [
        event["id"] 
        for event in all_events_for_tourney 
        if all(keyword not in event["name"].lower() for keyword in exclude_keywords) 
        and any(keyword in event["name"].lower() for keyword in include_keywords)
    ]
    if len(all_selections) > 1:
        logging.info(f"Uncertain which event is Ultimate Singles: {str(all_selections)}")
    else:
        return all_selections[0] if all_selections else None


def count_total_objects_in_set(set_data):
    # Count all objects in a single set
    count = 1  # 1 for the set itself
    count += 2  # 2 slot objects
    count += 2  # 2 entrant objects
    count += 2  # 2 standing objects
    count += 2  # 2 stats objects
    count += 2  # 2 score objects

    return count


# GraphQL query template
query = """
query EventSets($eventId: ID!, $page: Int!, $perPage: Int!) {
  event(id: $eventId) {
    id
    name
    startAt
    sets(
      page: $page
      perPage: $perPage
      sortType: STANDARD
    ) {
      pageInfo {
        total
      }
      nodes {
        slots {
          entrant {
            name
            participants{
                player{
                    user{
                        discriminator
                    }
                }
            }
          }
          standing {
            placement
            stats {
              score {
                value
              }
            }
          }
        }
      }
    }
  }
}
"""

# Function to extract player data from the set
def process_set_data(set_data):
    try:
        player_1 = set_data['slots'][0]['entrant']['name']
    except:
        player_1 = None
    try:
        id_1 = set_data['slots'][0]['entrant']['participants'][0]["player"]["user"]["discriminator"]
    except:
        id_1 = None
    try:
        score_1 = set_data['slots'][0]['standing']['stats']['score']['value']
    except:
        score_1 = None
    try:
        player_2 = set_data['slots'][1]['entrant']['name']
    except:
        player_2 = None
    try:
        id_2 = set_data['slots'][1]['entrant']['participants'][0]["player"]["user"]["discriminator"]
    except:
        id_2 = None
    try:
        score_2 = set_data['slots'][1]['standing']['stats']['score']['value']
    except:
        score_2 = None
    # Determine the winner based on the placement field
    if id_1 == "":
        id_1 = None
    if id_2 == "":
        id_2 = None
    try:
        if set_data['slots'][0]['standing']['placement'] < set_data['slots'][1]['standing']['placement']:
            winner_id = id_1
        else:
            winner_id = id_2
    except:
        winner_id = None

    # Create the desired dictionary format
    return {
        "player_1": player_1,
        "id_1": id_1,
        "score_1": score_1,
        "player_2": player_2,
        "id_2": id_2,
        "score_2": score_2,
        "winner_id": winner_id
    }

@limits(calls=80, period=60)
@on_exception(expo, RateLimitException, max_tries=5)
def rate_limited_request_post(url, json, headers):
    return requests.post(url, json=json, headers=headers)

@on_exception(expo, Exception, max_tries=5)
@on_exception(expo, ValueError, max_tries=5)
def get_all_sets(event_id, token):
    url = "https://api.start.gg/gql/alpha"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }

    all_sets = []
    page = 1
    per_page = 40  # Adjust as needed

    while True:
        variables = {
            "eventId": event_id,
            "page": page,
            "perPage": per_page
        }

        response = rate_limited_request_post(url, json={'query': query, 'variables': variables}, headers=headers)
        if response.status_code != 200:
            logging.info(variables)
            raise Exception(f"get_all_sets Request failed with status {response.status_code}")

        data = response.json()
        if not data.get("success", True) or "data" not in data:
            raise ValueError(f"get_all_sets No success or no data in response: {str(response)}")

        sets = data['data']['event']['sets']['nodes']
        start_time = datetime.utcfromtimestamp(data['data']['event']['startAt'])

        total_objects = 0
        for node_idx, set_data in enumerate(sets):
            if len(set_data['slots']) < 2:
                logging.info(f"Fewer than 2 slots in node {node_idx}:")
                logging.info(variables)
                continue
            processed_set = process_set_data(set_data)
            total_objects += count_total_objects_in_set(set_data)
            all_sets.append(processed_set)

        total_sets = data['data']['event']['sets']['pageInfo']['total']
        retrieved_sets = len(all_sets)

        logging.info(f"Total objects processed: {total_objects}")
        logging.info(f"Retrieved {retrieved_sets} out of {total_sets} sets")
        if retrieved_sets >= total_sets:
            break

        page += 1  # Move to the next page

    return start_time, all_sets


# Example usage:
def get_all_info_for_tournament(tourney_slug, event_slug):
    tournament_name, all_events_for_tourney = get_tournament_events(tourney_slug, token)
    if event_slug is not None:
        overall_slug = f"tournament/{tourney_slug}/event/{event_slug}"
        selected_event_id = get_event_id(overall_slug, token)
    else:
        selected_event_id = select_correct_event(all_events_for_tourney)
    event_start_time, sets = get_all_sets(selected_event_id, token)
    return tournament_name, event_start_time, sets


def extract_actual_value(cell_value):
    """
    Extracts the actual value from an IFERROR formula cell, or returns
    the value directly if it is not a formula.
    """
    if isinstance(cell_value, str) and cell_value.startswith('=IFERROR'):
        # Use regex to extract the value within double quotes, which represents the actual value
        match = re.search(r',"([^"]*)"\)', cell_value)
        if match:
            return match.group(1)  # Return the extracted value
        else:
            return None  # If no match, return None (or consider setting a default)
    return cell_value  # Return the value as-is if it is not wrapped in a formula


def process_tournament_file(file_path):
    # Define possible column names and the standardized names we want
    column_map = {
        'tournament': ['Tournament', 'Tournaments', 'Tournament*'],
        'link': ['Link', 'Links'],
        'tier': ['Category', 'Tier']
    }
    
    # Load Excel file using openpyxl if the file is .xlsx or .xls
    if file_path.endswith(('.xlsx', '.xls')):
        workbook = load_workbook(file_path, data_only=False)  # Do not use data_only to get formulas
        sheet = workbook[workbook.sheetnames[0]]  # Load the first sheet by default
        
        # Extract data from the sheet row by row
        data = []
        for row in sheet.iter_rows(min_row=2):
            row_data = {}
            tier_found = False  # Flag to mark if the first "Tier" column has been found
            #print("################")
            for cell in row:
                # Map cell content to appropriate column based on its position
                column_name = sheet.cell(row=1, column=cell.column).value
                #print(cell.value)
                if column_name in column_map['tournament']:
                    # Extract display name and hyperlink URL if available
                    row_data['tournament'] = extract_actual_value(cell.value)  # Tournament name
                    #print(cell.hyperlink.target)
                    row_data['link'] = cell.hyperlink.target if cell.hyperlink else None
                elif column_name in column_map['tier'] and not tier_found:
                    # Only set tier if it hasn't been set already
                    row_data['tier'] = extract_actual_value(cell.value)
                    tier_found = True  # Mark that the first "Tier" has been processed
                elif column_name in column_map['link'] and 'link' not in row_data:
                    row_data['link'] = extract_actual_value(cell.value)
            #print(row_data)
            
            # Append row data if 'tournament' column is populated (skip empty rows)
            if 'tournament' in row_data:
                data.append(row_data)
        # Convert extracted data to DataFrame
        df = pd.DataFrame(data)
        
        if 'link' not in df.columns:
            df['link'] = [row.get('link', None) for row in data]
    
    elif file_path.endswith('.csv'):
        # For CSV files, use pandas directly
        df = pd.read_csv(file_path)
        
        # Rename columns to standardized names using the mapping
        rename_columns = {}
        for standard_name, variations in column_map.items():
            for col in variations:
                if col in df.columns:
                    rename_columns[col] = standard_name
                    break  # Use the first matching column only
        df = df.rename(columns=rename_columns)
        
        # If 'link' column is missing, add a blank column for consistency
        if 'link' not in df.columns:
            df['link'] = None
    
    else:
        raise ValueError("Unsupported file format")
    
    # Final clean-up to ensure standardized column names and order
    table = df[['tournament', 'link', 'tier']]
    # Initialize a list to collect rows
    rows = []
    # Process each row in the table
    for _, row in table.iterrows():
        link = row["link"]

        # Skip rows without valid links
        if '.gg/' not in str(link) or "lne.gg" in str(link):
            if row["tournament"]:
                logging.info(f"BAD TOURNAMENT NO LINK: {str(row)}")
            continue

        # Extract components from the link
        try:
            link_parts = link.split('.gg/')[1].split('/')
            tourney_slug = link_parts[1]
            event_slug = link_parts[3]
        except IndexError:
            raise ValueError(f"Unexpected link format in {file_path}: {link}")

        # Clean and assign the tier information
        tier = str(row["tier"]).replace('Category', '').strip()
        if tier in tier_mapper:
            tier = tier_mapper[tier]

        # Append the row to the list
        rows.append([link, tourney_slug, event_slug, tier])

    print(len(rows))
    # Create a DataFrame from the rows list
    return pd.DataFrame(rows, columns=['Link', 'tourney_slug', 'event_slug', 'Tier'])


def get_s3_filenames(bucket_name, prefix=''):
    """
    Get the list of filenames in an S3 bucket.

    Parameters:
        bucket_name (str): The name of the S3 bucket.
        prefix (str): Optional prefix to filter files in a specific folder path.

    Returns:
        list: A list of filenames in the specified S3 bucket.
    """
    filenames = []

    # Paginate through all files if there are many
    paginator = s3.get_paginator('list_objects_v2')
    for page in paginator.paginate(Bucket=bucket_name, Prefix=prefix):
        if 'Contents' in page:
            for obj in page['Contents']:
                filenames.append(obj['Key'])
    
    return filenames


def get_major_tournaments_from_folder(directory_path):
    """
    Processes all tournament files in a directory, consolidating them into a single DataFrame.
    
    Parameters:
        directory_path (str): Path to the directory containing CSV files.
        
    Returns:
        pd.DataFrame: DataFrame containing consolidated data from all files in the directory.
    """
    # Initialize an empty DataFrame to store all tournament data
    all_tournaments_df = pd.DataFrame(columns=['Link', 'tourney_slug', 'event_slug', 'Tier'])

    # Process each file in the directory
    for filename in os.listdir(directory_path):
        file_path = os.path.join(directory_path, filename)
        logging.info(file_path)
        try:
            # Process the file and concatenate it to the main DataFrame
            tournament_df = process_tournament_file(file_path)
            all_tournaments_df = pd.concat([all_tournaments_df, tournament_df], ignore_index=True)
        except ValueError as e:
            logging.info(f"Error processing {file_path}: {e}")
        #break

    return all_tournaments_df


def download_google_sheet_as_excel(google_sheets_url, sheet_name, output_file):
    """
    Downloads a specific Google Sheets tab as an Excel file while preserving hyperlinks.
    Removes all other tabs except the specified one if there are multiple.

    Parameters:
        google_sheets_url (str): The URL of the Google Sheets document.
        sheet_name (str): The name of the tab to keep.
        output_file (str): The path to save the resulting Excel file.
    """
    # Convert the Google Sheets URL to the exportable Excel format
    excel_url = google_sheets_url.replace('/edit?gid=', '/export?format=xlsx&gid=')
    logging.info(excel_url)
    
    # Download the Excel file directly from Google Sheets
    response = requests.get(excel_url)
    response.raise_for_status()  # Check if the request was successful
    
    # Write the downloaded content to an Excel file
    with open(output_file, 'wb') as file:
        file.write(response.content)
    
    # Open the Excel file and remove other sheets except the specified one
    wb = load_workbook(output_file)
    logging.info(wb.sheetnames)
    for tab in wb.sheetnames:
        if tab != sheet_name:
            del wb[tab]
    
    # Save the updated Excel file with only the specified tab
    wb.save(output_file)
    logging.info(f"Downloaded and saved '{sheet_name}' tab to {output_file}.")



def bucket_exists(bucket_name):
    try:
        s3.head_bucket(Bucket=bucket_name)
        return True
    except ClientError:
        return False

def create_bucket(bucket_name):
    if not bucket_exists(bucket_name):
        s3.create_bucket(Bucket=bucket_name)
        logging.info(f"S3 bucket '{bucket_name}' created.")

def table_exists(table_name):
    try:
        dynamodb.Table(table_name).load()
        return True
    except ClientError:
        return False

def create_table(table_name):
    if not table_exists(table_name):
        table = dynamodb.create_table(
            TableName=table_name,
            KeySchema=[{'AttributeName': 'tourney_slug', 'KeyType': 'HASH'}],
            AttributeDefinitions=[{'AttributeName': 'tourney_slug', 'AttributeType': 'S'}],
            ProvisionedThroughput={'ReadCapacityUnits': 5, 'WriteCapacityUnits': 5}
        )
        table.wait_until_exists()
        logging.info(f"DynamoDB table '{table_name}' created.")

def upload_to_s3(bucket_name, file_path, key):
    s3.upload_file(file_path, bucket_name, key)
    logging.info(f"Uploaded {file_path} to {bucket_name}/{key}")

def update_dynamodb(table_name, item_data):
    table = dynamodb.Table(table_name)
    table.put_item(Item=item_data)
    logging.info(f"Updated DynamoDB with {item_data['tourney_slug']}")



def check_ddb_for_event(table_name, tourney_slug, event_slug):
    table = dynamodb.Table(table_name)

    try:
        response = table.query(
            KeyConditionExpression=Key('tourney_slug').eq(tourney_slug),
            FilterExpression=Key('event_slug').eq(event_slug),
            Limit=1  # Limit results to 1 for faster performance if it exists
        )
        # Check if any items were returned
        return len(response.get('Items', [])) > 0

    except ClientError as e:
        logging.info(f"Error querying DynamoDB: {e}")
        return False

def process_tournaments(google_sheets_url=None, sheet_name=None, tournament_folder_path=None, excluded_tiers=("D", "C", "B", "B+", "A", "A+")):
    # Ensure S3 bucket and DynamoDB table exist
    create_bucket(s3_bucket)
    logging.info(f"Checked S3 bucket '{s3_bucket}', created if it didn't exist.")
    
    create_table(dynamo_db_table_name)
    logging.info(f"Checked DynamoDB table '{dynamo_db_table_name}', created if it didn't exist.")

    # Load tournaments data from Google Sheets or folder path
    if google_sheets_url:
        logging.info(f"Downloading Google Sheet from {google_sheets_url}.")
        output_file = 'output_tmp.xlsx'
        download_google_sheet_as_excel(google_sheets_url, sheet_name, output_file)
        all_tournaments_df = process_tournament_file(output_file)
        logging.info(f"Processed tournament file from Google Sheets into DataFrame with {len(all_tournaments_df)} records.")
    elif tournament_folder_path:
        logging.info(f"Loading tournaments from folder {tournament_folder_path}.")
        all_tournaments_df = get_major_tournaments_from_folder(tournament_folder_path)
        logging.info(f"Loaded tournament data from folder into DataFrame with {len(all_tournaments_df)} records.")
    else:
        raise ValueError("Either Google Sheets URL or folder path must be provided.")

    # Get the list of existing files in S3 bucket
    s3_files = get_s3_filenames(s3_bucket)
    logging.info(f"Retrieved {len(s3_files)} files from S3 bucket '{s3_bucket}'.")

    total_tournaments_to_process = 0
    logging.info("Need to process:")
    for tourney_slug, event_slug, tier in zip(all_tournaments_df["tourney_slug"], all_tournaments_df["event_slug"],
                                              all_tournaments_df["Tier"]):
        json_filename = f"{tourney_slug}-{event_slug}.json"
        if json_filename not in s3_files and (tier not in excluded_tiers) and not check_ddb_for_event(dynamo_db_table_name, tourney_slug, event_slug):
            logging.info(tourney_slug)
            total_tournaments_to_process += 1
    logging.info(f"Total tournaments:{total_tournaments_to_process}")
    # Process each tournament entry
    processed_idx = 0
    for tourney_slug, event_slug, tier in zip(all_tournaments_df["tourney_slug"], all_tournaments_df["event_slug"], all_tournaments_df["Tier"]):
        json_filename = f"{tourney_slug}-{event_slug}.json"
        if (json_filename not in s3_files) and (tier not in excluded_tiers) and not check_ddb_for_event(dynamo_db_table_name, tourney_slug, event_slug):
            logging.info(f"Processing tournament '{tourney_slug}' with event '{event_slug}' and tier '{tier}'.")
            logging.info(f"{processed_idx}/{total_tournaments_to_process}")
            processed_idx += 1
            # Retrieve tournament details
            try:
                tournament_name, event_start_time, sets = get_all_info_for_tournament(tourney_slug, event_slug)
                logging.info(f"Retrieved details for tournament '{tourney_slug}': Name={tournament_name}, Date={event_start_time}, Sets={len(sets)}.")
            except Exception as e:
                logging.error(f"Failed to retrieve details for tournament '{tourney_slug}': {e}")
                continue

            # Prepare and save JSON data
            data = {
                "link": google_sheets_url,
                "event": event_slug,
                "tier": tier,
                "date": event_start_time.isoformat(),
                "name": tournament_name,
                "sets": sets
            }
            try:
                with open(json_filename, 'w') as f:
                    json.dump(data, f)
                logging.info(f"Saved tournament data for '{tourney_slug}' to file '{json_filename}'.")
            except Exception as e:
                logging.error(f"Failed to save JSON file '{json_filename}': {e}")
                continue

            # Upload JSON file to S3 and delete local copy
            try:
                upload_to_s3(s3_bucket, json_filename, json_filename)
                os.remove(json_filename)
                logging.info(f"Uploaded '{json_filename}' to S3 bucket '{s3_bucket}' and deleted local file.")
            except Exception as e:
                logging.error(f"Failed to upload '{json_filename}' to S3 or delete local file: {e}")
                continue

            # Update DynamoDB with tournament metadata
            dynamo_item = {
                "tourney_slug": tourney_slug,
                "event_slug": event_slug,
                "tier": tier,
                "date": event_start_time.isoformat(),
                "name": tournament_name
            }
            try:
                update_dynamodb(dynamo_db_table_name, dynamo_item)
                logging.info(f"Updated DynamoDB table '{dynamo_db_table_name}' with tournament '{tourney_slug}'.")
            except Exception as e:
                logging.error(f"Failed to update DynamoDB for '{tourney_slug}': {e}")









